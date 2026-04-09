import os
import threading
import time
import random
import urllib.parse
import re
import csv
import requests
import io
from datetime import datetime
from flask import Flask, request, jsonify, render_template
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

app = Flask(__name__)

# --- CONFIGURATION ---
PROFILE_PATH = r"C:\selenium-profile"
DEFAULT_COUNTRY_CODE = "91"

status_data = {
    "sent": 0,
    "failed": 0,
    "skipped": 0,
    "total": 0,
    "running": False,
    "scheduled_for": None,
    "aborted": False # New flag to handle thread termination
}

# [setup_driver, normalize_number, is_valid_number remain the same]

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-data-dir={PROFILE_PATH}")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

def normalize_number(num):
    num = re.sub(r"\D", "", str(num))
    if len(num) == 10: num = DEFAULT_COUNTRY_CODE + num
    return num

def is_valid_number(num):
    return re.fullmatch(r"\d{12,13}", num) is not None

def process_numbers(raw):
    global status_data
    seen = set()
    clean = []
    skipped_count = 0
    for r in raw:
        num = normalize_number(r.get("number"))
        name = r.get("name", "User")
        if not is_valid_number(num) or num in seen:
            skipped_count += 1
            continue
        seen.add(num)
        clean.append({"name": name, "number": num})
    status_data["skipped"] = skipped_count
    return clean

# [read_excel, read_csv, read_google_sheet remain the same]

def read_excel(path):
    try:
        wb = load_workbook(path)
        sheet = wb.active
        raw = [{"name": r[0] or "User", "number": r[1]} for r in sheet.iter_rows(min_row=2, values_only=True) if r[1]]
        return process_numbers(raw)
    except: return []

# ---------------- SENDER ENGINE ---------------- #

def send_bulk(data, message, config, schedule_delay=0):
    global status_data
    status_data["aborted"] = False

    if schedule_delay > 0:
        # Check for abort during the sleep period
        end_time = time.time() + schedule_delay
        while time.time() < end_time:
            if status_data["aborted"]:
                status_data["scheduled_for"] = None
                return
            time.sleep(1)

    status_data["sent"] = 0
    status_data["failed"] = 0
    status_data["total"] = len(data)
    status_data["running"] = True
    status_data["scheduled_for"] = None

    driver = None
    try:
        driver = setup_driver()
        driver.get("https://web.whatsapp.com")
        wait = WebDriverWait(driver, 60)
        wait.until(EC.presence_of_element_located((By.ID, "pane-side")))

        for i, user in enumerate(data):
            if status_data["aborted"] or not status_data["running"]: break
            
            phone = user['number']
            msg = message.replace("{name}", user['name'])
            
            try:
                url = f"https://web.whatsapp.com/send?phone={phone}&text={urllib.parse.quote(msg)}"
                driver.get(url)
                msg_box = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')))
                time.sleep(2)
                msg_box.send_keys(Keys.ENTER)
                status_data["sent"] += 1
            except:
                status_data["failed"] += 1

            time.sleep(random.randint(config["DELAY_MIN"], config["DELAY_MAX"]))
            
            if (i + 1) % config["BATCH"] == 0 and (i + 1) < len(data):
                cooldown = random.randint(config["COOL_MIN"], config["COOL_MAX"])
                end_cool = time.time() + cooldown
                while time.time() < end_cool:
                    if status_data["aborted"]: break
                    time.sleep(1)
    finally:
        if driver: driver.quit()
        status_data["running"] = False
        status_data["aborted"] = False

# ---------------- ROUTES ---------------- #

@app.route('/')
def index():
    return render_template("index14.html")

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    f = request.files['file']
    f.save("data.xlsx")
    contacts = read_excel("data.xlsx")
    return jsonify({"count": len(contacts)})

@app.route('/send', methods=['POST'])
def send():
    req = request.json
    config = {
        "DELAY_MIN": int(req.get("dmin", 10)),
        "DELAY_MAX": int(req.get("dmax", 25)),
        "BATCH": int(req.get("batch", 5)),
        "COOL_MIN": int(req.get("cmin", 2)) * 60,
        "COOL_MAX": 120
    }

    if req.get("numbers"):
        raw = [{"name": "User", "number": n.strip()} for n in req.get("numbers").replace('\n', ',').split(",") if n.strip()]
        data = process_numbers(raw)
    else:
        data = read_excel("data.xlsx")

    if not data:
        return jsonify({"error": "No valid data found. Import a file or enter numbers."}), 400

    schedule_delay = 0
    sched_time_str = req.get("schedule_time")
    if sched_time_str:
        sched_time = datetime.strptime(sched_time_str, "%Y-%m-%dT%H:%M")
        now = datetime.now()
        if sched_time > now:
            schedule_delay = (sched_time - now).total_seconds()
            status_data["scheduled_for"] = sched_time_str

    threading.Thread(target=send_bulk, args=(data, req.get("message"), config, schedule_delay)).start()
    return jsonify({"total": len(data), "scheduled": schedule_delay > 0})

@app.route('/abort', methods=['POST'])
def abort():
    global status_data
    status_data["aborted"] = True
    status_data["running"] = False
    status_data["scheduled_for"] = None
    return jsonify({"status": "stopping"})

@app.route('/status')
def status():
    return jsonify(status_data)

if __name__ == '__main__':
    app.run(debug=True)