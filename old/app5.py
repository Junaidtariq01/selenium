# 🚀 FULL WEB APP (v2): Excel Upload + Personalization + Scheduler + Dedup + Validation

# ================= BACKEND (Flask) =================

from flask import Flask, request, jsonify, render_template
import threading, time, random, urllib.parse, re, datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

app = Flask(__name__)

PROFILE_PATH = r"C:\selenium-profile"

# Anti-spam configs
MESSAGE_DELAY = (8, 20)
BATCH_SIZE = 5
BATCH_COOLDOWN = (60, 120)

status_data = {
    "sent": 0,
    "failed": 0,
    "skipped": 0,
    "running": False,
    "scheduled": False
}

# ---------------- DRIVER ---------------- #

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-data-dir={PROFILE_PATH}")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

# ---------------- HELPERS ---------------- #

def random_delay(a, b):
    time.sleep(random.randint(a, b))


def is_valid_number(num):
    return re.fullmatch(r"\d{10,13}", str(num)) is not None


def read_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    data = []
    seen = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, number = row[0], str(row[1])

        if number in seen:
            status_data["skipped"] += 1
            continue

        if not is_valid_number(number):
            status_data["skipped"] += 1
            continue

        seen.add(number)
        data.append({"name": name or "User", "number": number})

    return data

# ---------------- SENDER ---------------- #

def send_bulk(data, message_template):
    global status_data

    driver = setup_driver()
    driver.get("https://web.whatsapp.com")

    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.ID, "pane-side"))
    )

    status_data["running"] = True

    for i, user in enumerate(data):
        try:
            msg = message_template.replace("{name}", user['name'])
            encoded = urllib.parse.quote(msg)

            url = f"https://web.whatsapp.com/send?phone={user['number']}&text={encoded}"
            driver.get(url)

            box = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]'))
            )

            time.sleep(random.uniform(1, 3))
            box.send_keys(Keys.ENTER)

            status_data["sent"] += 1

        except Exception as e:
            status_data["failed"] += 1

        random_delay(*MESSAGE_DELAY)

        if (i + 1) % BATCH_SIZE == 0:
            time.sleep(random.randint(*BATCH_COOLDOWN))

    driver.quit()
    status_data["running"] = False

# ---------------- SCHEDULER ---------------- #

def schedule_task(run_time, data, message):
    status_data["scheduled"] = True

    while True:
        now = datetime.datetime.now()
        if now >= run_time:
            send_bulk(data, message)
            status_data["scheduled"] = False
            break
        time.sleep(5)

# ---------------- ROUTES ---------------- #

@app.route('/')
def index():
    return render_template("index5.html")

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    file_path = "data.xlsx"
    file.save(file_path)

    data = read_excel(file_path)

    return jsonify({"count": len(data)})

@app.route('/send', methods=['POST'])
def send():
    req = request.json
    data = read_excel("data.xlsx")
    message = req.get("message")

    schedule_time = req.get("schedule")

    if schedule_time:
        run_time = datetime.datetime.strptime(schedule_time, "%Y-%m-%d %H:%M")
        threading.Thread(target=schedule_task, args=(run_time, data, message)).start()
        return jsonify({"status": "scheduled"})

    threading.Thread(target=send_bulk, args=(data, message)).start()
    return jsonify({"status": "started"})

@app.route('/status')
def status():
    return jsonify(status_data)

# ================= FRONTEND =================


# ================= RUN =================

if __name__ == '__main__':
    app.run(debug=True)
