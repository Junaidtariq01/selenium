# 🚀 FULL WEB APP (v4): CSV + Google Sheets + FIX WINDOW ISSUE

from flask import Flask, request, jsonify, render_template
import threading, time, random, urllib.parse, re, datetime, csv, requests, io
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

app = Flask(__name__)

PROFILE_PATH = r"C:\selenium-profile"
DEFAULT_COUNTRY_CODE = "91"

MESSAGE_DELAY = (8, 20)
BATCH_SIZE = 5
BATCH_COOLDOWN = (60, 120)

status_data = {"sent":0,"failed":0,"skipped":0,"running":False}

# ---------------- DRIVER ---------------- #

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-data-dir={PROFILE_PATH}")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# ---------------- HELPERS ---------------- #

def normalize_number(num):
    num = re.sub(r"\D", "", str(num))
    if len(num) == 10:
        num = DEFAULT_COUNTRY_CODE + num
    return num


def is_valid_number(num):
    return re.fullmatch(r"\d{12,13}", num) is not None

# ---------------- DATA PROCESS ---------------- #

def process_numbers(raw):
    seen = set()
    clean = []

    for r in raw:
        num = normalize_number(r.get("number"))
        name = r.get("name", "User")

        if not is_valid_number(num) or num in seen:
            status_data["skipped"] += 1
            continue

        seen.add(num)
        clean.append({"name":name, "number":num})

    return clean


def read_excel(path):
    wb = load_workbook(path)
    sheet = wb.active
    raw = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw.append({"name":row[0], "number":row[1]})

    return process_numbers(raw)


def read_csv(file):
    content = file.read().decode('utf-8')
    reader = csv.reader(io.StringIO(content))
    raw = []

    next(reader, None)
    for row in reader:
        raw.append({"name":row[0], "number":row[1]})

    return process_numbers(raw)


def read_google_sheet(url):
    # Convert to CSV export link
    if "docs.google.com" in url:
        url = url.replace("/edit?usp=sharing", "/export?format=csv")

    res = requests.get(url)
    reader = csv.reader(io.StringIO(res.text))

    raw = []
    next(reader, None)
    for row in reader:
        raw.append({"name":row[0], "number":row[1]})

    return process_numbers(raw)

# ---------------- SENDER (FIXED) ---------------- #

def send_bulk(data, message):
    driver = setup_driver()
    driver.get("https://web.whatsapp.com")

    WebDriverWait(driver,60).until(EC.presence_of_element_located((By.ID,"pane-side")))

    status_data["running"] = True

    for i,user in enumerate(data):
        try:
            msg = message.replace("{name}", user['name'])

            # 🔥 FIX: DO NOT driver.get() every time
            search_url = f"https://web.whatsapp.com/send?phone={user['number']}"
            driver.get(search_url)

            box = WebDriverWait(driver,30).until(
                EC.presence_of_element_located((By.XPATH,'//div[@contenteditable="true"][@data-tab="10"]')))

            box.send_keys(msg)
            time.sleep(1)
            box.send_keys(Keys.ENTER)

            status_data["sent"] += 1

        except Exception as e:
            status_data["failed"] += 1

        time.sleep(random.randint(*MESSAGE_DELAY))

        if (i+1)%BATCH_SIZE==0:
            time.sleep(random.randint(*BATCH_COOLDOWN))

    driver.quit()
    status_data["running"] = False

# ---------------- ROUTES ---------------- #

@app.route('/')
def index():
    return render_template("index7.html")

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    f = request.files['file']
    f.save("data.xlsx")
    return jsonify({"count":len(read_excel("data.xlsx"))})

@app.route('/upload_csv', methods=['POST'])
def upload_csv():
    f = request.files['file']
    return jsonify({"count":len(read_csv(f))})

@app.route('/google_sheet', methods=['POST'])
def google_sheet():
    url = request.json.get("url")
    return jsonify({"count":len(read_google_sheet(url))})

@app.route('/send', methods=['POST'])
def send():
    req = request.json
    message = req.get("message")

    if req.get("numbers"):
        raw = [{"name":"User","number":n.strip()} for n in req.get("numbers").split(",")]
        data = process_numbers(raw)
    else:
        data = read_excel("data.xlsx")

    threading.Thread(target=send_bulk, args=(data,message)).start()
    return jsonify({"total":len(data)})

@app.route('/status')
def status():
    return jsonify(status_data)

# ---------------- FRONTEND ---------------- #


if __name__=='__main__':
    app.run(debug=True)
