# 🚀 FINAL ALL-IN-ONE SYSTEM (ULTIMATE VERSION)
# Excel + CSV + Google Sheets + Manual Input + Search Optimization + Dashboard Controls
import os
from flask import Flask, request, jsonify, render_template
import threading, time, random, urllib.parse, re, csv, requests, io
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

app = Flask(__name__)

PROFILE_PATH = r"C:\selenium-profile"
DEFAULT_COUNTRY_CODE = "91"

CONFIG = {
    "DELAY_MIN": 8,
    "DELAY_MAX": 20,
    "BATCH": 5,
    "COOL_MIN": 60,
    "COOL_MAX": 120
}

status_data = {"sent":0,"failed":0,"skipped":0,"running":False}

# ---------------- DRIVER ---------------- #

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-data-dir={PROFILE_PATH}")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# ---------------- HELPERS ---------------- #

def normalize_number(num):
    num = re.sub(r"\D", "", str(num))
    if len(num) == 10:
        num = DEFAULT_COUNTRY_CODE + num
    return num


def is_valid_number(num):
    return re.fullmatch(r"\d{12,13}", num) is not None


def process_numbers(raw):
    seen = set()
    clean = []

    for r in raw:
        num = normalize_number(r.get("number"))
        name = r.get("name", "User")

        if not is_valid_number(num) or num in seen:
            status_data["skipped"] += 1
            continue

        seen.add(num)
        clean.append({"name":name, "number":num})

    return clean

# ---------------- DATA SOURCES ---------------- #

def read_excel(path):
    wb = load_workbook(path)
    sheet = wb.active
    raw = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw.append({"name":row[0], "number":row[1]})
    return process_numbers(raw)


def read_csv(file):
    content = file.read().decode('utf-8')
    reader = csv.reader(io.StringIO(content))
    raw = []
    next(reader, None)
    for row in reader:
        raw.append({"name":row[0], "number":row[1]})
    return process_numbers(raw)


def read_google_sheet(url):
    if "docs.google.com" in url:
        url = url.replace("/edit?usp=sharing", "/export?format=csv")

    res = requests.get(url)
    reader = csv.reader(io.StringIO(res.text))

    raw = []
    next(reader, None)
    for row in reader:
        raw.append({"name":row[0], "number":row[1]})

    return process_numbers(raw)

# ---------------- SENDER (OPTIMIZED) ---------------- #



# def send_bulk(data, message):
    global status_data

    # reset status
    status_data["sent"] = 0
    status_data["failed"] = 0
    status_data["skipped"] = 0
    status_data["running"] = True

    driver = setup_driver()
    driver.get(f"https://web.whatsapp.com/send?phone={number}")

    wait = WebDriverWait(driver, 60)

    # wait for login
    wait.until(EC.presence_of_element_located((By.ID, "pane-side")))

    print("✅ WhatsApp Loaded")

    for i, user in enumerate(data):
        try:
            msg = message.replace("{name}", user['name'])

            # 🔥 SEARCH BOX (FIXED)
            search_box = wait.until(
                EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@title="Search input textbox"]'))
            )

            # clear properly (CTRL + A + DELETE)
            search_box.click()
            search_box.send_keys(Keys.CONTROL + "a")
            search_box.send_keys(Keys.DELETE)

            time.sleep(1)

            search_box.send_keys(user['number'])
            time.sleep(2)
            search_box.send_keys(Keys.ENTER)

            print(f"Opened chat: {user['number']}")

            # 🔥 MESSAGE BOX (FIXED XPATH)
            msg_box = wait.until(
                EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]'))
            )

            msg_box.click()
            msg_box.send_keys(msg)
            time.sleep(1)
            msg_box.send_keys(Keys.ENTER)

            print(f"Sent to {user['number']}")
            status_data["sent"] += 1

        except Exception as e:
            print(f"❌ Failed: {user['number']} | Error: {e}")
            status_data["failed"] += 1

        # delay
        delay = random.randint(CONFIG["DELAY_MIN"], CONFIG["DELAY_MAX"])
        time.sleep(delay)

        # batch cooldown
        if (i + 1) % CONFIG["BATCH"] == 0:
            cooldown = random.randint(CONFIG["COOL_MIN"], CONFIG["COOL_MAX"])
            print(f"⏸ Cooling for {cooldown}s")
            time.sleep(cooldown)

    driver.quit()
    status_data["running"] = False
def send_bulk(data, message):
    global status_data

    status_data.update({"sent": 0, "failed": 0, "skipped": 0, "running": True})

    driver = setup_driver()
    driver.get("https://web.whatsapp.com")

    wait = WebDriverWait(driver, 60)

    # wait for login
    wait.until(EC.presence_of_element_located((By.ID, "pane-side")))
    print("✅ WhatsApp Loaded")

    for i, user in enumerate(data):
        try:
            msg = message.replace("{name}", user['name'])

            phone = user['number']
            url = f"https://web.whatsapp.com/send?phone={phone}&text={urllib.parse.quote(msg)}"

            # 🔥 open chat WITHOUT restarting browser
            driver.get(url)

            # wait for message box
            msg_box = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')
                )
            )

            time.sleep(2)

            # press enter to send
            msg_box.send_keys(Keys.ENTER)

            print(f"✅ Sent to {phone}")
            status_data["sent"] += 1

        except Exception as e:
            print(f"❌ Failed: {phone} | Error: {e}")
            status_data["failed"] += 1

        # delay
        delay = random.randint(CONFIG["DELAY_MIN"], CONFIG["DELAY_MAX"])
        time.sleep(delay)

        # batch cooldown
        if (i + 1) % CONFIG["BATCH"] == 0:
            cooldown = random.randint(CONFIG["COOL_MIN"], CONFIG["COOL_MAX"])
            print(f"⏸ Cooling for {cooldown}s")
            time.sleep(cooldown)

    driver.quit()
    status_data["running"] = False
# ---------------- ROUTES ---------------- #

@app.route('/')
def index():
    return render_template("index9.html")

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    f = request.files['file']
    f.save("data.xlsx")
    return jsonify({"count":len(read_excel("data.xlsx"))})

@app.route('/upload_csv', methods=['POST'])
def upload_csv():
    f = request.files['file']
    return jsonify({"count":len(read_csv(f))})

@app.route('/google_sheet', methods=['POST'])
def google_sheet():
    url = request.json.get("url")
    return jsonify({"count":len(read_google_sheet(url))})

@app.route('/send', methods=['POST'])
def send():
    req = request.json

    CONFIG["DELAY_MIN"] = int(req.get("dmin",8))
    CONFIG["DELAY_MAX"] = int(req.get("dmax",20))
    CONFIG["BATCH"] = int(req.get("batch",5))
    CONFIG["COOL_MIN"] = int(req.get("cmin",60))
    CONFIG["COOL_MAX"] = int(req.get("cmax",120))

    if req.get("numbers"):
        raw = [{"name":"User","number":n.strip()} for n in req.get("numbers").split(",")]
        data = process_numbers(raw)
    else:
        data = read_excel("data.xlsx")

    threading.Thread(target=send_bulk, args=(data,req.get("message"))).start()
    return jsonify({"total":len(data)})

@app.route('/status')
def status():
    return jsonify(status_data)

# ---------------- FRONTEND ---------------- #


if __name__=='__main__':
    app.run(debug=True)
