# 🚀 FULL WEB APP (v3 HYBRID): Excel + Manual Input + Smart Dedup (with/without country code)

from flask import Flask, request, jsonify, render_template
import threading, time, random, urllib.parse, re, datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

app = Flask(__name__)

PROFILE_PATH = r"C:\selenium-profile"
DEFAULT_COUNTRY_CODE = "91"

MESSAGE_DELAY = (8, 20)
BATCH_SIZE = 5
BATCH_COOLDOWN = (60, 120)

status_data = {"sent":0,"failed":0,"skipped":0,"running":False}

# ---------------- DRIVER ---------------- #

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-data-dir={PROFILE_PATH}")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# ---------------- HELPERS ---------------- #

def normalize_number(num):
    num = re.sub(r"\D", "", str(num))  # remove non-digits

    if len(num) == 10:
        num = DEFAULT_COUNTRY_CODE + num

    return num


def is_valid_number(num):
    return re.fullmatch(r"\d{12,13}", num) is not None

# ---------------- DATA PROCESSING ---------------- #

def process_numbers(raw_list):
    seen = set()
    clean_data = []

    for entry in raw_list:
        name = entry.get("name", "User")
        num = normalize_number(entry.get("number"))

        if not is_valid_number(num):
            status_data["skipped"] += 1
            continue

        if num in seen:
            status_data["skipped"] += 1
            continue

        seen.add(num)
        clean_data.append({"name": name, "number": num})

    return clean_data


def read_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    raw = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw.append({"name": row[0], "number": row[1]})

    return process_numbers(raw)

# ---------------- SENDER ---------------- #

def send_bulk(data, message_template):
    global status_data
    driver = setup_driver()
    driver.get("https://web.whatsapp.com")

    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "pane-side")))
    status_data["running"] = True

    for i, user in enumerate(data):
        try:
            msg = message_template.replace("{name}", user['name'])
            encoded = urllib.parse.quote(msg)
            driver.get(f"https://web.whatsapp.com/send?phone={user['number']}&text={encoded}")

            box = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')))

            time.sleep(random.uniform(1,3))
            box.send_keys(Keys.ENTER)
            status_data["sent"] += 1

        except:
            status_data["failed"] += 1

        time.sleep(random.randint(*MESSAGE_DELAY))

        if (i+1) % BATCH_SIZE == 0:
            time.sleep(random.randint(*BATCH_COOLDOWN))

    driver.quit()
    status_data["running"] = False

# ---------------- ROUTES ---------------- #

@app.route('/')
def index():
    return render_template("index6.html")

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    file.save("data.xlsx")
    data = read_excel("data.xlsx")
    return jsonify({"count": len(data)})

@app.route('/send', methods=['POST'])
def send():
    req = request.json
    message = req.get("message")

    # HYBRID INPUT
    if req.get("numbers"):
        raw_numbers = req.get("numbers").split(",")
        raw = [{"name": "User", "number": n.strip()} for n in raw_numbers]
        data = process_numbers(raw)
    else:
        data = read_excel("data.xlsx")

    threading.Thread(target=send_bulk, args=(data, message)).start()
    return jsonify({"status":"started","total":len(data)})

@app.route('/status')
def status():
    return jsonify(status_data)

# ---------------- FRONTEND ---------------- #



if __name__ == '__main__':
    app.run(debug=True)
