from flask import Flask, request, jsonify, render_template
import threading, time, random, urllib.parse, re, csv, requests, io, os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

app = Flask(__name__)

# --- CONFIGURATION ---
PROFILE_PATH = r"C:\selenium-profile"
DEFAULT_COUNTRY_CODE = "91"

# Centralized State
status_data = {
    "sent": 0, 
    "failed": 0, 
    "skipped": 0, 
    "running": False, 
    "total": 0  # Added for the progress bar
}

# This will hold the numbers in memory so you don't have to keep reading files
pending_data = []

# ---------------- HELPERS ---------------- #

def normalize_number(num):
    num = re.sub(r"\D", "", str(num))
    if len(num) == 10:
        num = DEFAULT_COUNTRY_CODE + num
    return num

def is_valid_number(num):
    return re.fullmatch(r"\d{12,13}", num) is not None

def process_raw_data(raw_list):
    """Standardizes data from any source into a clean list"""
    global status_data
    seen = set()
    clean = []
    skipped_count = 0

    for r in raw_list:
        name = str(r.get("name") or "User").strip()
        num = normalize_number(r.get("number"))

        if not is_valid_number(num) or num in seen:
            skipped_count += 1
            continue

        seen.add(num)
        clean.append({"name": name, "number": num})
    
    status_data["skipped"] = skipped_count
    return clean

# ---------------- DRIVER SETUP ---------------- #

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-data-dir={PROFILE_PATH}")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--start-maximized")
    # Uncomment below to run without a visible window after testing
    # options.add_argument("--headless") 

    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

# ---------------- SENDER ENGINE ---------------- #

# def send_bulk(data, message, config):
    global status_data
    driver = None
    
    try:
        status_data["running"] = True
        status_data["total"] = len(data)
        status_data["sent"] = 0
        status_data["failed"] = 0

        driver = setup_driver()
        driver.get("https://web.whatsapp.com")
        
        # Wait for WhatsApp to load (Check for the chat list sidebar)
        wait = WebDriverWait(driver, 60)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-testid='chat-list']")))

        for i, user in enumerate(data):
            if not status_data["running"]: break # Ability to stop if needed

            try:
                msg = message.replace("{name}", user['name'])
                phone = user['number']
                encoded_msg = urllib.parse.quote(msg)
                
                # Direct API link injection
                driver.get(f"https://web.whatsapp.com/send?phone={phone}&text={encoded_msg}")

                # Wait for the send button to be clickable
                send_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@data-icon="send"]')))
                time.sleep(1) # Small human-like buffer
                send_btn.click()
                
                status_data["sent"] += 1
                print(f"✅ [{i+1}/{len(data)}] Sent to {phone}")

            except Exception as e:
                status_data["failed"] += 1
                print(f"❌ Failed {user['number']}: {str(e)[:50]}")

            # Smart Delay
            time.sleep(random.randint(config["dmin"], config["dmax"]))

            # Batch Cooldown
            if (i + 1) % config["batch"] == 0 and (i + 1) != len(data):
                cooldown = random.randint(config["cmin"], config["cmax"])
                print(f"⏸ Batch complete. Cooling down for {cooldown}s...")
                time.sleep(cooldown)

    except Exception as global_e:
        print(f"🚨 Critical System Error: {global_e}")
    finally:
        if driver:
            driver.quit()
        status_data["running"] = False

def send_bulk(data, message, config):
    global status_data
    driver = None
    
    try:
        status_data["running"] = True
        status_data["total"] = len(data)
        driver = setup_driver()
        wait = WebDriverWait(driver, 30) # 30s is usually enough for elements

        # Initial Login Check
        driver.get("https://web.whatsapp.com")
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-testid='chat-list']")))

        for i, user in enumerate(data):
            if not status_data["running"]: break 

            try:
                phone = user['number']
                msg = message.replace("{name}", user['name'])
                url = f"https://web.whatsapp.com/send?phone={phone}&text={urllib.parse.quote(msg)}"
                
                driver.get(url)

                # 1. Check for "Invalid Number" popup immediately
                try:
                    invalid_popup = WebDriverWait(driver, 8).until(
                        EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'invalid')]"))
                    )
                    print(f"⚠️ Invalid number: {phone}")
                    status_data["failed"] += 1
                    continue # Skip to next contact
                except:
                    pass # Number is likely valid, proceed to send

                # 2. Wait for the message box OR the send button
                # We use ENTER on the message box as it's often more stable than the icon
                msg_box = wait.until(EC.presence_of_element_located(
                    (By.XPATH, '//div[@title="Type a message" or @contenteditable="true"]')
                ))
                
                time.sleep(1) # Human-like pause
                msg_box.send_keys(Keys.ENTER)
                
                status_data["sent"] += 1
                print(f"✅ [{i+1}/{len(data)}] Sent to {phone}")

            except Exception as e:
                status_data["failed"] += 1
                print(f"❌ Error with {user['number']}: {str(e)[:50]}")

            time.sleep(random.randint(config["dmin"], config["dmax"]))

            # Batch cooldown logic...
            if (i + 1) % config["batch"] == 0 and (i + 1) != len(data):
                time.sleep(random.randint(config["cmin"], config["cmax"]))

    finally:
        if driver: driver.quit()
        status_data["running"] = False
# ---------------- ROUTES ---------------- #

@app.route('/')
def index():
    return render_template("index11g.html")

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    global pending_data
    f = request.files['file']
    wb = load_workbook(f)
    sheet = wb.active
    raw = [{"name": row[0], "number": row[1]} for row in sheet.iter_rows(min_row=2, values_only=True)]
    pending_data = process_raw_data(raw)
    return jsonify({"count": len(pending_data)})

@app.route('/upload_csv', methods=['POST'])
def upload_csv():
    global pending_data
    f = request.files['file']
    content = f.read().decode('utf-8')
    reader = csv.reader(io.StringIO(content))
    next(reader, None) # skip header
    raw = [{"name": row[0], "number": row[1]} for row in reader if len(row) >= 2]
    pending_data = process_raw_data(raw)
    return jsonify({"count": len(pending_data)})

@app.route('/google_sheet', methods=['POST'])
def google_sheet():
    global pending_data
    url = request.json.get("url")
    if "docs.google.com" in url:
        url = url.split('/edit')[0] + "/export?format=csv"
    
    res = requests.get(url)
    reader = csv.reader(io.StringIO(res.text))
    next(reader, None)
    raw = [{"name": row[0], "number": row[1]} for row in reader if len(row) >= 2]
    pending_data = process_raw_data(raw)
    return jsonify({"count": len(pending_data)})

@app.route('/send', methods=['POST'])
def start_send():
    global pending_data, status_data
    req = request.json

    # 1. Grab numbers (Manual Entry priority, then File Uploads)
    if req.get("numbers"):
        raw = [{"name": "User", "number": n.strip()} for n in req.get("numbers").replace('\n', ',').split(",") if n.strip()]
        data_to_send = process_raw_data(raw)
    else:
        data_to_send = pending_data

    if not data_to_send:
        return jsonify({"error": "No numbers found"}), 400

    # 2. Extract Config
    config = {
        "dmin": int(req.get("dmin", 10)),
        "dmax": int(req.get("dmax", 25)),
        "batch": int(req.get("batch", 5)),
        "cmin": int(req.get("cmin", 60)),
        "cmax": int(req.get("cmax", 120))
    }

    # 3. Reset Status and Start Thread
    status_data.update({"sent": 0, "failed": 0, "skipped": 0, "total": len(data_to_send)})
    threading.Thread(target=send_bulk, args=(data_to_send, req.get("message"), config)).start()
    
    return jsonify({"total": len(data_to_send)})

@app.route('/status')
def get_status():
    return jsonify(status_data)

if __name__ == '__main__':
    app.run(debug=True, port=5000)