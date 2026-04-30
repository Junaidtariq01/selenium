import os
import threading
import time
import random
import re
import csv
from wsgiref import headers
import pyautogui
from datetime import datetime
from flask import request, jsonify
from flask_login import login_required
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# --- CONFIGURATION ---
PROFILE_PATH = r"C:\selenium-profile"
DEFAULT_COUNTRY_CODE = "91"
TEMP_FILE = "data.xlsx"
ATTACHMENT_FILE = "None"
sendTimeFlag = False

status_data = {
    "sent": 0, "failed": 0, "skipped": 0, "total": 0,
    "aborted": False, "running": False, "paused": False,
    "scheduled_for": None, "countdown": 0, "current_action": "Idle",
    "typing_speed": 0,"show_notice": False
}

# ---------------- HELPERS ---------------- #

def smart_sleep(seconds, action_text="Waiting"):
    global status_data
    status_data["countdown"] = int(seconds)
    status_data["current_action"] = action_text
    while status_data["countdown"] > 0:
        if status_data["aborted"]:
            status_data["countdown"] = 0
            return
        while status_data["paused"]:
            if status_data["aborted"]: return
            status_data["current_action"] = "Paused"
            time.sleep(0.5)
        time.sleep(1)
        status_data["countdown"] -= 1
    status_data["countdown"] = 0

def normalize_number(num):
    if not num: return ""
    num = str(num).strip()
    if num.startswith('+'): num = num[1:]
    num = re.sub(r"\D", "", num) # Remove spaces, dashes, dots
    if num.startswith('00'): num = num[2:]
    if len(num) == 11 and num.startswith('0'): num = num[1:]
    if len(num) == 10: num = DEFAULT_COUNTRY_CODE + num
    return num

def is_valid_number(num):
    return re.fullmatch(r"\d{11,15}", num) is not None

def process_numbers(raw):
    global status_data
    seen, clean = set(), []
    skipped_count = 0
    for r in raw:
        num = normalize_number(r.get("number"))
        if not is_valid_number(num) or num in seen:
            skipped_count += 1
            continue
        seen.add(num)
        clean.append({
            "name": r.get("name", "User"), 
            "number": num,
            "custom_msg": r.get("custom_msg")
            })
    status_data["skipped"] = skipped_count
    return clean

def is_safe_hour():
    hour = datetime.now().hour
    return 9 <= hour < 21

def spin_message(text):
    pattern = r'\{([^{}|]+\|[^{}]+)\}'
    while True:
        match = re.search(pattern, text)
        if not match: break
        options = match.group(1).split('|')
        text = text.replace(match.group(0), random.choice(options), 1)
    return text



def upload_attachment():
    global ATTACHMENT_FILE
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files['file']

    if f.filename == '':
        return jsonify({"error": "Empty file"}), 400

    path = os.path.join(os.getcwd(), f.filename)
    f.save(path)
    print("📥 Saved attachment at:", path)
    ATTACHMENT_FILE = path

    return jsonify({"status": "uploaded"})



# ---------------- SENDER ENGINE ---------------- #


def send_bulk(data, message, config, schedule_delay=0):
    global status_data, ATTACHMENT_FILE, sendTimeFlag

    status_data.update({
        "aborted": False,
        "paused": False,
        "sent": 0,
        "failed": 0,
        "total": len(data),
        "running": True
    })

    if schedule_delay > 0:
        smart_sleep(schedule_delay, "Scheduling")

    driver = None

    try:

        options = webdriver.ChromeOptions()
        options.add_argument(f"user-data-dir={PROFILE_PATH}")
        options.add_argument("--remote-debugging-port=9222")

        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )

        driver.get("https://web.whatsapp.com")
        wait = WebDriverWait(driver, 60)

        status_data["current_action"] = "Waiting for WhatsApp Login..."
        wait.until(EC.presence_of_element_located((By.ID, "pane-side")))

        for i, user in enumerate(data):

            if status_data["aborted"]:
                break

            while status_data["paused"]:
                if status_data["aborted"]:
                    break
                time.sleep(1)

            if config.get("SAFE_HOURS") and not is_safe_hour():
                smart_sleep(900, "Outside Safe Hours")
                continue

            phone = user['number']
            # msg = message.replace("{name}", user['name'])

            # Priority Logic: Use Excel message if present, else UI template
            custom = user.get("custom_msg")
            if custom and custom.strip().lower() != "none": # Check for "None" string from Excel[cite: 7]
                msg = custom
            else:
                msg = message.replace("{name}", user['name'])

            # SAFEGUARD: Ensure msg is a string to prevent NoneType error[cite: 7]
            msg = str(msg) if msg is not None else ""

            if config.get("SPINTAX") and msg:
                msg = spin_message(msg)


            # if user.get("custom_msg"):
            #     msg = user["custom_msg"]
            # else:
            #     msg = message.replace("{name}", user['name'])

            # if config.get("SPINTAX"):
            #     msg = spin_message(msg)

            try:
                status_data["current_action"] = f"Opening chat of {phone}"
                driver.get(f"https://web.whatsapp.com/send?phone={phone}")

                # Wait for chat OR invalid popup
                try:
                    element = WebDriverWait(driver, 15).until(
                        EC.presence_of_element_located((
                            By.XPATH,
                            '//div[@contenteditable="true"][@data-tab="10"] | //div[contains(text(),"invalid")] | //div[contains(text(),"not on WhatsApp")]'
                        ))
                    )

                    if "invalid" in element.text.lower() or "not on whatsapp" in element.text.lower():
                        status_data["skipped"] += 1
                        status_data["current_action"] = f"Skipped: {phone}"
                        continue

                    msg_box = element

                except:
                    status_data["failed"] += 1
                    continue

                # =========================
                # 📎 ATTACHMENT SECTION
                # =========================

                # if ATTACHMENT_FILE and os.path.exists(ATTACHMENT_FILE):
                #     YOUR_CLIP_X = 669  # 🔥 X coordinate of attach button
                #     YOUR_CLIP_Y = 994  # 🔥 Y coordinate of attach button
                #     YOUR_PHOTO_X = 663  # 🔥 X coordinate of "Photos & Videos" button
                #     YOUR_PHOTO_Y = 727  # 🔥  Y coordinate of "Photos & Videos" button


                #     try:
                #         abs_path = os.path.abspath(ATTACHMENT_FILE)
                #         status_data["current_action"] = "Sending Attachment..."
                #         print("📎 Sending file:", abs_path)

                #         driver.maximize_window()
                #         time.sleep(5)

                #         # STEP 1: Click attach button
                #         pyautogui.moveTo(YOUR_CLIP_X, YOUR_CLIP_Y, duration=2.5)
                #         pyautogui.click()
                #         time.sleep(5)

                #         # STEP 2: Click "Photos & Videos"
                #         pyautogui.moveTo(YOUR_PHOTO_X, YOUR_PHOTO_Y, duration=1)
                #         pyautogui.click()

                #         # 🔥 IMPORTANT WAIT (dialog must open)
                #         time.sleep(5)

                #         # STEP 3: Now type path (dialog is active)
                #         interval = random.uniform(0.05, 0.45)
                #         pyautogui.write(abs_path, interval=interval)
                #         time.sleep(3)

                #         pyautogui.press("enter")  # select file

                #         # STEP 4: wait preview
                #         time.sleep(3)

                #         # STEP 5: send
                #         pyautogui.press("enter")
                #         time.sleep(5)

                #         print("✅ Attachment sent successfully")

                #         sendTimeFlag = True
                #     except Exception as e:
                #         print("❌ PyAutoGUI error:", e)
                #         status_data["failed"] += 1
                #         continue
                

                if ATTACHMENT_FILE and os.path.exists(ATTACHMENT_FILE):
                    try:
                        abs_path = os.path.abspath(ATTACHMENT_FILE).replace("\\", "/")

                        print("📎 Sending file:", abs_path)
                        print("📁 EXISTS:", os.path.exists(abs_path))

                        wait.until(EC.presence_of_element_located((By.XPATH, '//footer')))
                        time.sleep(2)

                        file_inputs = driver.find_elements(By.XPATH, '//input[@type="file"]')

                        if not file_inputs:
                            raise Exception("❌ No file input found")

                        uploaded = False

                        for inp in file_inputs:
                            try:
                                accept = inp.get_attribute("accept")
                                print("🔎 Input accept:", accept)

                                if accept and ("image" in accept or "video" in accept):

                                    driver.execute_script("""
                                        arguments[0].style.display = 'block';
                                        arguments[0].style.visibility = 'visible';
                                        arguments[0].style.opacity = 1;
                                    """, inp)

                                    inp.send_keys(abs_path)
                                    uploaded = True
                                    print("✅ Correct input used")
                                    break

                            except:
                                continue

                        if not uploaded:
                            raise Exception("❌ No valid file input matched")

                        # wait for preview
                        time.sleep(5)

                        # find send button
                        # send_buttons = driver.find_elements(By.XPATH, '//span[@data-icon="send"]')

                        # print("📤 Found send buttons:", len(send_buttons))

                        # if not send_buttons:
                        #     raise Exception("❌ Send button not found (preview failed)")

                        # driver.execute_script("arguments[0].click();", send_buttons[-1])

                         # STEP 5: send
                        pyautogui.press("enter")
                        time.sleep(5)

                        print("✅ Attachment sent successfully")

                        sendTimeFlag = True

                    except Exception as e:
                        print("❌ Attachment error:", e)
                        status_data["failed"] += 1
                        continue

                # =========================
                # ✉️ TEXT MESSAGE SECTION
                # =========================
                if msg.strip():

                    if sendTimeFlag == True:
                        time.sleep(15)
                  

                    status_data["current_action"] = f"Typing to {phone}"

                    if config.get("TYPING"):
                        for char in msg:
                            if status_data["aborted"]:
                                break

                            speed = random.uniform(0.03, 0.14)
                            status_data["typing_speed"] = round(speed * 1000, 0)

                            msg_box.send_keys(char)
                            time.sleep(speed)

                        status_data["typing_speed"] = 0
                    else:
                        msg_box.send_keys(msg)

                    time.sleep(1.5)
                    msg_box.send_keys(Keys.ENTER)
                    time.sleep(10)

                status_data["sent"] += 1

            except Exception as e:
                print(f"Error sending to {phone}: {e}")
                status_data["failed"] += 1

            # =========================
            # ⏱ DELAYS & BATCH
            # =========================
            if i < len(data) - 1:
                delay = random.randint(config["DELAY_MIN"], config["DELAY_MAX"])

                if (i + 1) % config["BATCH"] == 0:
                    delay = random.randint(config["COOL_MIN"], config["COOL_MAX"])
                    smart_sleep(delay, "Batch Cooldown")
                else:
                    smart_sleep(delay, "Next Delay")

    finally:
        if driver:
            driver.quit()

        status_data.update({
            "running": False,
            "current_action": "Completed",
            "countdown": 0
        })

        if status_data["sent"] > 0 and os.path.exists(TEMP_FILE):
            try:
                os.remove(TEMP_FILE)
            except:
                pass

        if ATTACHMENT_FILE and os.path.exists(ATTACHMENT_FILE):
            try:
                os.remove(ATTACHMENT_FILE)
            except:
                pass

# ---------------- API ENDPOINTS ---------------- #

@login_required
def get_status():
    return jsonify(status_data)

@login_required
def upload_excel():
    f = request.files['file']
    f.save(TEMP_FILE)
    wb = load_workbook(TEMP_FILE)
    sheet = wb.active

    # Dynamically find the index of the "Messages" column
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]]
    msg_idx = headers.index("messages") if "messages" in headers else -1
    

    raw = []
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if r[1]: # If number exists
            raw.append({
                "name": str(r[0] or "User"), 
                "number": str(r[1]),
                "custom_msg": str(r[msg_idx]) if msg_idx != -1 and r[msg_idx] else None
            })
    data = process_numbers(raw)

    # raw = [{"name": str(r[0] or "User"), "number": str(r[1])} for r in sheet.iter_rows(min_row=2, values_only=True) if r[1]]
    # data = process_numbers(raw)
    return jsonify({"count": len(data)})

@login_required
def upload_csv():
    f = request.files['file']
    f.save("data.csv")
    raw = []

    # with open("data.csv", mode='r', encoding='utf-8') as infile:
        # reader = csv.DictReader(infile)
        # for row in reader:
            # raw.append({"name": row.get("name", "User"), "number": row.get("number")})
    
    
    # ... previous code ...
    with open("data.csv", mode='r', encoding='utf-8') as infile:
        reader = csv.DictReader(infile)
        for row in reader:
            # Capture "Messages" or "messages" column if present
            raw.append({
                "name": row.get("name", "User"), 
                "number": row.get("number"),
                "custom_msg": row.get("Messages") or row.get("messages")
            })   
    
    
    data = process_numbers(raw)
    return jsonify({"count": len(data)})

@login_required
def start_campaign():
    req = request.json
    manual_input = req.get("numbers", "").strip()
    
    # Priority 1: Manual Input
    if manual_input:
        # Split by comma or newline
        raw_list = re.split(r'[,\n]+', manual_input)
        raw = [{"name": "User", "number": n.strip()} for n in raw_list if n.strip()]
        data = process_numbers(raw)
    # Priority 2: Excel File
    elif os.path.exists(TEMP_FILE):
        wb = load_workbook(TEMP_FILE)
        sheet = wb.active

        # ADD THIS: Look for the Messages column here too
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]]
        msg_idx = headers.index("messages") if "messages" in headers else -1

        raw = []
        for r in sheet.iter_rows(min_row=2, values_only=True):
            if r[1]: 
                raw.append({
                    "name": str(r[0] or "User"), 
                    "number": str(r[1]),
                    "custom_msg": str(r[msg_idx]) if msg_idx != -1 and r[msg_idx] else None
                })
        data = process_numbers(raw)
        # raw = [{"name": str(r[0] or "User"), "number": str(r[1])} for r in sheet.iter_rows(min_row=2, values_only=True) if r[1]]
        # data = process_numbers(raw)
    else:
        return jsonify({"error": "No contacts found. Please upload file or enter numbers."}), 400

    if not data:
        return jsonify({"error": "No valid numbers after normalization."}), 400

    config = {
        "DELAY_MIN": int(req.get("dmin", 15)),
        "DELAY_MAX": int(req.get("dmax", 45)),
        "BATCH": int(req.get("batch", 5)),
        "COOL_MIN": int(req.get("cmin", 2)) * 60,
        "COOL_MAX": int(req.get("cmin", 2)) * 60 + 60,
        "TYPING": req.get("typing", True),
        "SPINTAX": req.get("spintax", True),
        "SAFE_HOURS": req.get("safe_hours", False)
    }

    schedule_delay = 0
    sched_time_str = req.get("schedule_time")
    if sched_time_str:
        try:
            sched_time = datetime.strptime(sched_time_str, "%Y-%m-%dT%H:%M")
            now = datetime.now()
            if sched_time > now:
                schedule_delay = (sched_time - now).total_seconds()
                status_data["scheduled_for"] = sched_time_str
        except: pass

    threading.Thread(target=send_bulk, args=(data, req.get("message"), config, schedule_delay)).start()
    return jsonify({"status": "started", "total": len(data)})

@login_required
def pause_campaign():
    status_data["paused"] = True
    return jsonify({"status": "paused"})

@login_required
def resume_campaign():
    status_data["paused"] = False
    return jsonify({"status": "resumed"})

@login_required
def abort_campaign():
    status_data.update({"aborted": True, "running": False, "paused": False, "scheduled_for": None})
    return jsonify({"status": "stopping"})