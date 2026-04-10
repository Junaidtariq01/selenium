import os
import threading
import time
import random
import urllib.parse
import re
import csv
import requests
import io
from datetime import datetime
from flask import Flask, request, jsonify, render_template
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

app = Flask(__name__)

# --- CONFIGURATION ---
PROFILE_PATH = r"C:\selenium-profile"
DEFAULT_COUNTRY_CODE = "91"

status_data = {
    "sent": 0,
    "failed": 0,
    "skipped": 0,
    "total": 0,
    "aborted": False,
    "running": False,
    "paused": False,
    "scheduled_for": None,
    "countdown": 0,
    "current_action": "Idle"
}

# ---------------- HELPERS ---------------- #

def smart_sleep(seconds, action_text="Waiting"):
    """Sleeps while allowing for pause/abort and updating countdown."""
    global status_data
    status_data["countdown"] = int(seconds)
    status_data["current_action"] = action_text
    
    while status_data["countdown"] > 0:
        if status_data["aborted"]:
            status_data["countdown"] = 0
            return
        
        # Pause Loop
        while status_data["paused"]:
            if status_data["aborted"]: return
            status_data["current_action"] = "Paused"
            time.sleep(0.5)
            
        status_data["current_action"] = action_text
        time.sleep(1)
        status_data["countdown"] -= 1
    
    status_data["countdown"] = 0

def normalize_number(num):
    """Advanced normalization for various formats."""
    if not num: return ""
    num = str(num).strip()
    # Remove leading +
    if num.startswith('+'): num = num[1:]
    # Remove all non-digits (spaces, dashes, etc.)
    num = re.sub(r"\D", "", num)
    # Handle double zero prefix
    if num.startswith('00'): num = num[2:]
    # Handle leading 0 (if 11 digits, likely 0 + 10 digit number)
    if len(num) == 11 and num.startswith('0'): num = num[1:]
    # Add default country code if exactly 10 digits
    if len(num) == 10: num = DEFAULT_COUNTRY_CODE + num
    return num

def is_valid_number(num):
    return re.fullmatch(r"\d{11,15}", num) is not None

def process_numbers(raw):
    global status_data
    seen = set()
    clean = []
    skipped_count = 0
    for r in raw:
        num = normalize_number(r.get("number"))
        name = r.get("name", "User")
        if not is_valid_number(num) or num in seen:
            skipped_count += 1
            continue
        seen.add(num)
        clean.append({"name": name, "number": num})
    status_data["skipped"] = skipped_count
    return clean

# [read_excel and other file helpers remain the same]
def read_excel(path):
    try:
        wb = load_workbook(path)
        sheet = wb.active
        raw = [{"name": r[0] or "User", "number": r[1]} for r in sheet.iter_rows(min_row=2, values_only=True) if r[1]]
        return process_numbers(raw)
    except: return []

def spin_message(text):
    pattern = r'\{([^{}|]+\|[^{}]+)\}'
    while True:
        match = re.search(pattern, text)
        if not match: break
        options = match.group(1).split('|')
        text = text.replace(match.group(0), random.choice(options), 1)
    return text

def is_safe_hour():
    hour = datetime.now().hour
    return 9 <= hour < 21

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument(f"user-data-dir={PROFILE_PATH}")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# ---------------- SENDER ENGINE ---------------- #

def send_bulk(data, message, config, schedule_delay=0):
    global status_data
    status_data["aborted"] = False
    status_data["paused"] = False

    if schedule_delay > 0:
        smart_sleep(schedule_delay, "Scheduling")

    status_data["sent"] = 0
    status_data["failed"] = 0
    status_data["total"] = len(data)
    status_data["running"] = True
    status_data["scheduled_for"] = None

    driver = None
    try:
        driver = setup_driver()
        driver.get("https://web.whatsapp.com")
        wait = WebDriverWait(driver, 60)
        status_data["current_action"] = "Waiting for WhatsApp Login..."
        wait.until(EC.presence_of_element_located((By.ID, "pane-side")))

        for i, user in enumerate(data):
            if status_data["aborted"]: break
            
            # Pause Check
            while status_data["paused"]:
                if status_data["aborted"]: break
                status_data["current_action"] = "Paused"
                time.sleep(1)

            if config.get("SAFE_HOURS") and not is_safe_hour():
                smart_sleep(900, "Outside Safe Hours")
                continue

            phone = user['number']
            msg = message.replace("{name}", user['name'])
            if config.get("SPINTAX"): msg = spin_message(msg)
            
            try:
                status_data["current_action"] = f"Sending to {phone}"
                driver.get(f"https://web.whatsapp.com/send?phone={phone}")
                msg_box = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')))
                
                if config.get("TYPING"):
                    for char in msg:
                        if status_data["aborted"]: break
                        msg_box.send_keys(char)
                        time.sleep(random.uniform(0.02, 0.08))
                else:
                    msg_box.send_keys(msg)
                
                time.sleep(1)
                msg_box.send_keys(Keys.ENTER)
                status_data["sent"] += 1
            except:
                status_data["failed"] += 1

            if random.random() > 0.7:
                driver.execute_script(f"window.scrollBy(0, {random.randint(100, 300)});")

            # Interval Delay
            if i < len(data) - 1:
                delay = random.randint(config["DELAY_MIN"], config["DELAY_MAX"])
                
                # Check for batch cooldown
                if (i + 1) % config["BATCH"] == 0:
                    delay = random.randint(config["COOL_MIN"], config["COOL_MAX"])
                    smart_sleep(delay, "Batch Cooldown")
                else:
                    smart_sleep(delay, "Next Message Delay")
                    
    finally:
        if driver: driver.quit()
        status_data["running"] = False
        status_data["current_action"] = "Completed"
        status_data["countdown"] = 0
        
    try:
        if status_data["sent"] > 0 and os.path.exists("data.xlsx"):
            os.remove("data.xlsx")
    except: pass

# ---------------- ROUTES ---------------- #

@app.route('/')
def index():
    return render_template("index3.html")

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    f = request.files['file']
    f.save("data.xlsx")
    contacts = read_excel("data.xlsx")
    return jsonify({"count": len(contacts)})

@app.route('/pause', methods=['POST'])
def pause():
    global status_data
    status_data["paused"] = True
    return jsonify({"status": "paused"})

@app.route('/resume', methods=['POST'])
def resume():
    global status_data
    status_data["paused"] = False
    return jsonify({"status": "resumed"})

@app.route('/abort', methods=['POST'])
def abort():
    global status_data
    status_data["aborted"] = True
    status_data["running"] = False
    status_data["paused"] = False
    status_data["scheduled_for"] = None
    return jsonify({"status": "stopping"})

@app.route('/send', methods=['POST'])
def send():
    req = request.json
    config = {
        "DELAY_MIN": int(req.get("dmin", 10)),
        "DELAY_MAX": int(req.get("dmax", 50)),
        "BATCH": int(req.get("batch", 5)),
        "COOL_MIN": int(req.get("cmin", 2)) * 60,
        "COOL_MAX": int(req.get("cmin", 2)) * 60 + 60,
        "TYPING": req.get("typing", True),
        "SPINTAX": req.get("spintax", True),
        "SAFE_HOURS": req.get("safe_hours", False)
    }

    if req.get("numbers"):
        raw = [{"name": "User", "number": n.strip()} for n in req.get("numbers").replace('\n', ',').split(",") if n.strip()]
        data = process_numbers(raw)
    else:
        data = read_excel("data.xlsx")

    if not data:
        return jsonify({"error": "No valid data found."}), 400

    schedule_delay = 0
    sched_time_str = req.get("schedule_time")
    if sched_time_str:
        try:
            sched_time = datetime.strptime(sched_time_str, "%Y-%m-%dT%H:%M")
            now = datetime.now()
            if sched_time > now:
                schedule_delay = (sched_time - now).total_seconds()
                status_data["scheduled_for"] = sched_time_str
        except: pass

    threading.Thread(target=send_bulk, args=(data, req.get("message"), config, schedule_delay)).start()
    return jsonify({"total": len(data), "scheduled": schedule_delay > 0})

@app.route('/status')
def status():
    return jsonify(status_data)

if __name__ == '__main__':
    app.run(debug=True, port=5000)